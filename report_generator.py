#!/usr/bin/env python3
"""
Report Generator Module
Handles HTML and Excel report generation, formatting, and file operations
"""

from datetime import datetime
from typing import List, Dict, Any, Optional
from config import config
from technical_analyzer import CandidateResult

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ReportGenerator:
    """Handles HTML and Excel report generation"""
    
    def __init__(self):
        self.config = config
        
    def generate_html_report(
        self, 
        panel_data: Dict[str, Any],
        portfolio_analysis: List[Dict[str, Any]],
        early_momentum: List[CandidateResult], 
        trend_candidates: List[CandidateResult],
        declining_funds: List[Dict[str, Any]],
        case_results: Dict[str, List],
        stats: Any
    ) -> str:
        """Generate comprehensive HTML report"""
        
        html_content = self._generate_html_header()
        html_content += self._generate_summary_section(panel_data, declining_funds, early_momentum, trend_candidates)
        html_content += self._generate_scoring_system_section()
        html_content += self._generate_portfolio_section(portfolio_analysis)
        html_content += self._generate_candidates_section(early_momentum, trend_candidates)
        html_content += self._generate_alerts_section(declining_funds)
        html_content += self._generate_case_results_section(case_results)
        html_content += self._generate_statistics_section(stats)
        html_content += self._generate_html_footer()
        
        # Save HTML file
        html_file = config.files.html_report_file
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"📄 HTML Rapor kaydedildi: {html_file}")
        print(f"🔍 Raporu görüntülemek için: open {html_file}")
        
        return html_file
    
    def _generate_html_header(self) -> str:
        """Generate HTML header with CSS styles"""
        return f"""
<!DOCTYPE html>
<html lang="tr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TEFAS Günlük Takip Paneli</title>
    <style>
        /* Modern CSS Reset & Base */
        * {{ box-sizing: border-box; }}
        body {{ 
            font-family: -apple-system, BlinkMacSystemFont, 'SF Pro Display', 'Helvetica Neue', Roboto, sans-serif; 
            margin: 0;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            background-attachment: fixed;
            min-height: 100vh;
            color: #2c3e50;
            line-height: 1.4;
            font-size: 13px;
        }}
        
        .container {{ 
            max-width: 1400px; 
            margin: 0 auto; 
            padding: 20px;
        }}
        
        /* Modern Header with Glassmorphism */
        .header {{ 
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(20px);
            border: 1px solid rgba(255, 255, 255, 0.2);
            color: #2c3e50;
            padding: 40px;
            border-radius: 24px;
            margin-bottom: 40px;
            text-align: center;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
            position: relative;
            overflow: hidden;
        }}
        
        .header::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
            background: linear-gradient(90deg, #667eea, #764ba2, #f093fb, #f5576c, #4facfe, #00f2fe);
        }}
        
        .header h1 {{ 
            font-size: 2.5em; 
            font-weight: 700; 
            margin: 0 0 10px 0;
            background: linear-gradient(135deg, #667eea, #764ba2);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}
        
        /* Enhanced Summary Cards */
        .summary {{ 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(280px, 1fr)); 
            gap: 24px; 
            margin-bottom: 40px; 
        }}
        
        .summary-card {{ 
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(20px);
            padding: 30px;
            border-radius: 20px;
            text-align: center;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
            border: 1px solid rgba(255, 255, 255, 0.2);
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
        }}
        
        .summary-card::before {{
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 4px;
            background: var(--card-gradient, linear-gradient(90deg, #667eea, #764ba2));
        }}
        
        .summary-card:hover {{ 
            transform: translateY(-8px) scale(1.02);
            box-shadow: 0 20px 40px rgba(0, 0, 0, 0.15);
        }}
        
        .summary-number {{ 
            font-size: 3.2em; 
            font-weight: 800; 
            margin-bottom: 8px;
            background: var(--number-gradient, linear-gradient(135deg, #667eea, #764ba2));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        }}
        
        .summary-card div:last-child {{ 
            font-weight: 600; 
            font-size: 1.1em;
            color: #555;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        
        /* Color Themes for Cards */
        .summary-card:nth-child(1) {{ --card-gradient: linear-gradient(90deg, #4facfe, #00f2fe); --number-gradient: linear-gradient(135deg, #4facfe, #00f2fe); }}
        .summary-card:nth-child(2) {{ --card-gradient: linear-gradient(90deg, #f093fb, #f5576c); --number-gradient: linear-gradient(135deg, #f093fb, #f5576c); }}
        .summary-card:nth-child(3) {{ --card-gradient: linear-gradient(90deg, #4facfe, #00f2fe); --number-gradient: linear-gradient(135deg, #4facfe, #00f2fe); }}
        .summary-card:nth-child(4) {{ --card-gradient: linear-gradient(90deg, #ff9a9e, #fecfef); --number-gradient: linear-gradient(135deg, #ff9a9e, #fecfef); }}
        
        .positive {{ color: #27ae60; }}
        .negative {{ color: #e74c3c; }}
        
        /* Modern Section Cards */
        .section {{ 
            background: rgba(255, 255, 255, 0.95);
            backdrop-filter: blur(20px);
            padding: 35px;
            margin-bottom: 30px;
            border-radius: 20px;
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1);
            border: 1px solid rgba(255, 255, 255, 0.2);
            position: relative;
        }}
        
        .section h2 {{ 
            margin: 0 0 25px 0;
            font-size: 1.8em;
            font-weight: 700;
            color: #2c3e50;
            padding-bottom: 10px;
            border-bottom: 3px solid;
            border-image: linear-gradient(90deg, #667eea, #764ba2) 1;
        }}
        
        /* Enhanced Fund Cards */
        .fund-card {{ 
            background: rgba(255, 255, 255, 0.7);
            border: 1px solid rgba(255, 255, 255, 0.3);
            border-radius: 16px;
            padding: 24px;
            margin-bottom: 20px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            backdrop-filter: blur(10px);
            position: relative;
            overflow: hidden;
        }}
        
        .fund-card::before {{
            content: '';
            position: absolute;
            left: 0;
            top: 0;
            bottom: 0;
            width: 5px;
            background: var(--trend-color, #3498db);
            border-radius: 0 8px 8px 0;
        }}
        
        .fund-card:hover {{ 
            transform: translateY(-4px);
            box-shadow: 0 12px 24px rgba(0, 0, 0, 0.15);
            background: rgba(255, 255, 255, 0.9);
        }}
        
        .trend-up {{ --trend-color: #27ae60; }}
        .trend-down {{ --trend-color: #e74c3c; }}
        .trend-neutral {{ --trend-color: #f39c12; }}
        
        /* Advanced Metrics Display */
        .stats-grid {{ 
            display: grid; 
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); 
            gap: 12px; 
            margin: 16px 0;
        }}
        
        .performance-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 12px;
            margin: 16px 0;
        }}
        
        .metric {{ 
            background: rgba(255, 255, 255, 0.5);
            padding: 8px 10px;
            border-radius: 10px;
            border: 1px solid rgba(255, 255, 255, 0.3);
            backdrop-filter: blur(10px);
            transition: all 0.2s ease;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            max-width: 100%;
        }}
        
        .metric:hover {{ 
            background: rgba(255, 255, 255, 0.8);
            transform: scale(1.02);
        }}
        
        .metric-label {{ 
            font-weight: 700;
            color: #7f8c8d;
            font-size: 0.7em;
            text-transform: uppercase;
            letter-spacing: 0.2px;
            display: block;
            margin-bottom: 2px;
            line-height: 1.2;
        }}
        
        .metric-value {{ 
            color: #2c3e50;
            font-weight: 600;
            font-size: 0.85em;
            line-height: 1.2;
            word-break: break-word;
        }}
        
        .performance-metric {{
            background: rgba(255, 255, 255, 0.6);
            padding: 8px 12px;
            border-radius: 10px;
            border: 1px solid rgba(255, 255, 255, 0.4);
            backdrop-filter: blur(10px);
            text-align: center;
            white-space: nowrap;
        }}
        
        .performance-label {{
            font-size: 0.65em;
            font-weight: 600;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 0.2px;
            margin-bottom: 1px;
            line-height: 1.2;
        }}
        
        .performance-value {{
            font-size: 0.8em;
            font-weight: 700;
            line-height: 1.2;
        }}
        
        .perf-positive {{ color: #27ae60; }}
        .perf-negative {{ color: #e74c3c; }}
        .perf-neutral {{ color: #6c757d; }}
        
        /* Enhanced Candidate Cards */
        .candidate-item {{ 
            background: rgba(255, 255, 255, 0.8);
            padding: 24px;
            margin-bottom: 16px;
            border-radius: 16px;
            backdrop-filter: blur(15px);
            border: 1px solid rgba(255, 255, 255, 0.3);
            position: relative;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            overflow: hidden;
        }}
        
        .candidate-item::before {{
            content: '';
            position: absolute;
            left: 0;
            top: 0;
            bottom: 0;
            width: 5px;
            background: var(--item-gradient, linear-gradient(180deg, #667eea, #764ba2));
        }}
        
        .candidate-item:hover {{ 
            transform: translateY(-4px) scale(1.01);
            box-shadow: 0 12px 24px rgba(0, 0, 0, 0.15);
            background: rgba(255, 255, 255, 0.95);
        }}
        
        .early-momentum {{ --item-gradient: linear-gradient(180deg, #f39c12, #e67e22); }}
        .strong-trend {{ --item-gradient: linear-gradient(180deg, #27ae60, #2ecc71); }}
        .decline-warning {{ --item-gradient: linear-gradient(180deg, #e74c3c, #c0392b); }}
        
        /* Ultra Modern TEFAS Button */
        .tefas-btn {{
            display: inline-flex;
            align-items: center;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            text-decoration: none;
            padding: 8px 16px;
            border-radius: 25px;
            font-size: 0.9em;
            font-weight: 600;
            margin-left: 12px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3);
            border: 1px solid rgba(255, 255, 255, 0.2);
            backdrop-filter: blur(10px);
            position: relative;
            overflow: hidden;
        }}
        
        .tefas-btn::before {{
            content: '';
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.2), transparent);
            transition: left 0.6s;
        }}
        
        .tefas-btn:hover {{
            background: linear-gradient(135deg, #5a67d8 0%, #667eea 100%);
            transform: translateY(-2px) scale(1.05);
            box-shadow: 0 8px 25px rgba(102, 126, 234, 0.4);
            color: white;
            text-decoration: none;
        }}
        
        .tefas-btn:hover::before {{
            left: 100%;
        }}
        
        .tefas-btn::after {{
            content: '🔗';
            margin-left: 6px;
            font-size: 0.9em;
        }}
        
        /* Advanced Case Analysis */
        .case-section {{ 
            margin-bottom: 24px;
            background: rgba(255, 255, 255, 0.5);
            padding: 20px;
            border-radius: 16px;
            backdrop-filter: blur(10px);
            border: 1px solid rgba(255, 255, 255, 0.3);
        }}
        
        .case-title {{ 
            color: #2c3e50;
            font-weight: 700;
            margin-bottom: 16px;
            font-size: 1.2em;
            padding: 12px 20px;
            background: rgba(102, 126, 234, 0.1);
            border-radius: 12px;
            border-left: 4px solid #667eea;
        }}
        
        /* Version Badge */
        .version-badge {{ 
            background: linear-gradient(135deg, #667eea, #764ba2);
            color: white;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 0.8em;
            font-weight: 600;
            box-shadow: 0 2px 8px rgba(102, 126, 234, 0.3);
            border: 1px solid rgba(255, 255, 255, 0.2);
        }}
        
        .timestamp {{ 
            opacity: 0.8;
            font-size: 0.9em;
            font-weight: 500;
        }}
        
        /* Responsive Design */
        @media (max-width: 768px) {{
            .container {{ padding: 10px; }}
            .header {{ padding: 20px; }}
            .header h1 {{ font-size: 1.8em; }}
            .summary {{ grid-template-columns: 1fr; gap: 16px; }}
            .summary-number {{ font-size: 2.5em; }}
            .stats-grid {{ grid-template-columns: 1fr; }}
            .tefas-btn {{ padding: 6px 12px; font-size: 0.8em; }}
        }}
        
        /* Smooth Animations */
        * {{ transition: all 0.2s ease; }}
        
        /* Custom Scrollbar */
        ::-webkit-scrollbar {{ width: 8px; }}
        ::-webkit-scrollbar-track {{ background: rgba(255, 255, 255, 0.1); border-radius: 10px; }}
        ::-webkit-scrollbar-thumb {{ background: rgba(102, 126, 234, 0.5); border-radius: 10px; }}
        ::-webkit-scrollbar-thumb:hover {{ background: rgba(102, 126, 234, 0.7); }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📈 TEFAS Günlük Takip Paneli <span class="version-badge">OOP v2.0</span></h1>
            <p class="timestamp">Güncelleme: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Sistem: Object-Oriented Architecture</p>
        </div>
"""
    
    def _generate_summary_section(self, panel_data: Dict, declining_funds: List, early_momentum: List, trend_candidates: List) -> str:
        """Generate summary section"""
        return f"""
        <div class="summary">
            <div class="summary-card">
                <div class="summary-number">{len(panel_data.get('portfolio_funds', []))}</div>
                <div>Portföy Fonları</div>
            </div>
            <div class="summary-card">
                <div class="summary-number positive">{len(early_momentum)}</div>
                <div>Erken Momentum</div>
            </div>
            <div class="summary-card">
                <div class="summary-number positive">{len(trend_candidates)}</div>
                <div>Güçlü Trend</div>
            </div>
            <div class="summary-card">
                <div class="summary-number negative">{len(declining_funds)}</div>
                <div>Aktif Uyarılar</div>
            </div>
        </div>
"""

    def _generate_scoring_system_section(self) -> str:
        """Generate comprehensive 100-point scoring system explanation"""
        return f"""
        <div class="section" style="background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%); border: 1px solid #dee2e6;">
            <h2>🎯 Puanlama Sistemi</h2>
            
            <div style="background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 15px; border-radius: 12px; margin-bottom: 25px; text-align: center;">
                <strong style="font-size: 1.2em;">📊 5 Ana Bileşenden Oluşan Profesyonel Analiz Sistemi</strong><br>
                <span style="opacity: 0.9; font-size: 0.9em;">Normalizasyonlu 100-Puan Skalası (Ham puan / 11 * 10)</span>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 25px; margin-bottom: 25px;">
                <!-- Trend Bileşeni -->
                <div style="background: rgba(39, 174, 96, 0.1); padding: 20px; border-radius: 12px; border-left: 4px solid #27ae60;">
                    <h3 style="color: #27ae60; margin-bottom: 15px; font-size: 1.1em;">📈 1. TREND BİLEŞENİ (40 Puan)</h3>
                    <ul style="margin: 0; padding-left: 18px; line-height: 1.7; font-size: 0.9em;">
                        <li><strong>RSI > 60:</strong> +{self.config.technical.rsi_strong_score} puan</li>
                        <li><strong>MACD Pozitif Kesişim:</strong> +{self.config.technical.macd_score} puan</li>
                        <li><strong>Bollinger Kırılımı:</strong> +{self.config.technical.bollinger_score} puan</li>
                    </ul>
                </div>
                
                <!-- Yatırımcı Akımı -->
                <div style="background: rgba(52, 152, 219, 0.1); padding: 20px; border-radius: 12px; border-left: 4px solid #3498db;">
                    <h3 style="color: #3498db; margin-bottom: 15px; font-size: 1.1em;">👥 2. YATIRIMCI AKIMI (20 Puan)</h3>
                    <ul style="margin: 0; padding-left: 18px; line-height: 1.7; font-size: 0.9em;">
                        <li><strong>7-gün yatırımcı artışı >%3</strong></li>
                        <li><strong>+ AUM (fon büyüklüğü) artışı</strong></li>
                        <li>Her iki koşul: +{self.config.technical.investor_flow_score} puan</li>
                    </ul>
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; margin-bottom: 25px;">
                <!-- Hacim & Likidite -->
                <div style="background: rgba(243, 156, 18, 0.1); padding: 18px; border-radius: 12px; border-left: 4px solid #f39c12;">
                    <h3 style="color: #f39c12; margin-bottom: 12px; font-size: 1.0em;">💰 3. HACİM & LİKİDİTE</h3>
                    <p style="margin: 0; font-size: 0.85em; line-height: 1.6;">
                        <strong>Günlük hacim artışı</strong><br>
                        <strong>+ 5-gün ort. yükseliş</strong><br>
                        <span style="color: #f39c12; font-weight: bold;">+{self.config.technical.volume_liquidity_score} puan</span>
                    </p>
                </div>
                
                <!-- Volatilite -->
                <div style="background: rgba(155, 89, 182, 0.1); padding: 18px; border-radius: 12px; border-left: 4px solid #9b59b6;">
                    <h3 style="color: #9b59b6; margin-bottom: 12px; font-size: 1.0em;">📊 4. VOLATİLİTE</h3>
                    <p style="margin: 0; font-size: 0.85em; line-height: 1.6;">
                        <strong>Düşük ATR artışı</strong><br>
                        <strong>+ Sharpe oranı >1</strong><br>
                        <span style="color: #9b59b6; font-weight: bold;">+{self.config.technical.volatility_score} puan</span>
                    </p>
                </div>
                
                <!-- Zamanlama -->
                <div style="background: rgba(231, 76, 60, 0.1); padding: 18px; border-radius: 12px; border-left: 4px solid #e74c3c;">
                    <h3 style="color: #e74c3c; margin-bottom: 12px; font-size: 1.0em;">⏰ 5. ZAMANLAMA</h3>
                    <p style="margin: 0; font-size: 0.85em; line-height: 1.6;">
                        <strong>Son 5 günde</strong><br>
                        <strong>4+ pozitif kapanış</strong><br>
                        <span style="color: #e74c3c; font-weight: bold;">+{self.config.technical.timing_score} puan</span>
                    </p>
                </div>
            </div>
            
            <!-- Skor Aralıkları -->
            <div style="background: linear-gradient(135deg, #ff9a9e 0%, #fecfef 100%); padding: 20px; border-radius: 12px; margin-top: 25px;">
                <h3 style="color: #2c3e50; margin-bottom: 15px; text-align: center;">🎯 ADAY KATEGORİLERİ</h3>
                <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                    <div style="background: rgba(255,255,255,0.8); padding: 15px; border-radius: 10px; text-align: center;">
                        <strong style="color: #f39c12; font-size: 1.1em;">🟡 ERKEN MOMENTUM</strong><br>
                        <span style="font-size: 1.3em; font-weight: bold; color: #2c3e50;">{self.config.analysis.early_momentum_min_score}-{self.config.analysis.early_momentum_max_score} Puan</span><br>
                        <small style="color: #6c757d;">Gelişmekte olan fırsatlar</small>
                    </div>
                    <div style="background: rgba(255,255,255,0.8); padding: 15px; border-radius: 10px; text-align: center;">
                        <strong style="color: #27ae60; font-size: 1.1em;">💎 GÜÇLÜ TREND</strong><br>
                        <span style="font-size: 1.3em; font-weight: bold; color: #2c3e50;">{self.config.analysis.trend_min_score}+ Puan</span><br>
                        <small style="color: #6c757d;">Yüksek güvenilirlik</small>
                    </div>
                </div>
            </div>
            
            <div style="background: #d4edda; padding: 12px; border-radius: 8px; margin-top: 20px; border-left: 4px solid #28a745; font-size: 0.9em;">
                <strong>💡 BİLGİ:</strong> Bu sistem temel teknik analizin ötesinde, yatırımcı davranışları, likidite durumu, 
                risk-getiri profili ve piyasa zamanlaması gibi <strong>profesyonel yatırım kriterlerini</strong> birleştirir.
            </div>
            
            <div style="background: #fff3cd; padding: 12px; border-radius: 8px; margin-top: 15px; border-left: 4px solid #ffc107; font-size: 0.85em;">
                <strong>🤹 NORMALIZASYON:</strong> Ham puanlama 110 üzerinden yapılır ve <strong>Skor = (Ham Puan ÷ 11) × 10</strong> 
                formülü ile 100-puan skalasına çevrilir. Bu sayede tüm bileşenler eşit ağırlıkta değerlendirilir.
            </div>
            
            <div style="background: #e7f3ff; padding: 12px; border-radius: 8px; margin-top: 15px; border-left: 4px solid #007bff; font-size: 0.85em;">
                <strong>🚀 MOMENTUM FİLTRESİ:</strong> Erken momentum ve güçlü trend adayları için ek filtre: 
                <strong>Son 3 günde en az 2 pozitif kapanış</strong> VEYA <strong>3-günlük ortalama değişim > %0.5</strong> koşulunu sağlamalı.
            </div>
            
            <div style="background: #f8f9fa; padding: 12px; border-radius: 8px; margin-top: 15px; border-left: 4px solid #6c757d; font-size: 0.85em;">
                <strong>📊 CASE ANALİZ FİLTRESİ:</strong> Özel pattern analizleri yalnızca <strong>yatırımcı sayısı 50 ve üstü</strong> 
                olan fonlar için yapılır. Bu sayede likidite riski azaltılır.
            </div>
        </div>
"""

    def _generate_portfolio_section(self, portfolio_analysis: List[Dict[str, Any]]) -> str:
        """Generate portfolio section with performance sorting"""
        if not portfolio_analysis:
            return """
        <div class="section">
            <h2>💼 Portföy Durumu</h2>
            <p>⚠️ Portföy fonları analiz edilemedi</p>
        </div>
"""
        
        # Sort by weekly performance (highest to lowest)
        sorted_portfolio = sorted(portfolio_analysis, 
                                key=lambda x: x.get('performance_7d', 0), 
                                reverse=True)
        
        portfolio_html = """
        <div class="section">
            <h2>💼 Portföy Durumu</h2>
            <div style="background: rgba(102, 126, 234, 0.1); padding: 12px; border-radius: 10px; margin-bottom: 20px; font-size: 0.85em;">
                📊 Fonlar haftalık kümülatif performansa göre sıralanmıştır (yüksekten düşüğe)
            </div>
"""
        
        for fund in sorted_portfolio:
            trend_class = "trend-up" if "🔥" in fund['Trend_Status'] or "📈" in fund['Trend_Status'] else \
                         "trend-down" if "🔴" in fund['Trend_Status'] or "📉" in fund['Trend_Status'] else "trend-neutral"
            
            investor_str = f"{fund['Yatırımcı']:,}" if isinstance(fund['Yatırımcı'], int) else str(fund['Yatırımcı'])
            
            # Get performance data
            perf_1d = fund.get('performance_1d', 0)
            perf_3d = fund.get('performance_3d', 0) 
            perf_7d = fund.get('performance_7d', 0)
            
            # Format performance with colors
            def format_perf(value):
                if value > 0:
                    return f"<span class='perf-positive'>+{value:.2f}%</span>"
                elif value < 0:
                    return f"<span class='perf-negative'>{value:.2f}%</span>"
                else:
                    return f"<span class='perf-neutral'>0.00%</span>"
            
            # Get comprehensive score from fund data
            comprehensive_score = fund.get('Comprehensive_Score', fund.get('Skor', 0))
            
            portfolio_html += f"""
            <div class="fund-card {trend_class}">
                <h3>{fund['Fon']} 
                    <span style="background: linear-gradient(135deg, #666, #888); color: white; padding: 2px 8px; border-radius: 12px; font-size: 0.7em; font-weight: normal; margin-left: 8px;">🏢 {fund.get('Şirket', 'Bilinmeyen')}</span>
                    <span style="background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 2px 8px; border-radius: 12px; font-size: 0.7em; font-weight: bold; margin-left: 8px;">SKOR: {comprehensive_score}</span>
                    <span style="color: #6c757d; font-size: 0.8em; margin-left: 8px;">| {fund['Trend_Status']}</span>
                    <a href="https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={fund['Fon']}" 
                       target="_blank" class="tefas-btn">TEFAS'ta Gör</a>
                </h3>
                
                <!-- Performance Metrics -->
                <div class="performance-grid">
                    <div class="performance-metric">
                        <div class="performance-label">1 Gün</div>
                        <div class="performance-value">{format_perf(perf_1d)}</div>
                    </div>
                    <div class="performance-metric">
                        <div class="performance-label">3 Gün</div>
                        <div class="performance-value">{format_perf(perf_3d)}</div>
                    </div>
                    <div class="performance-metric">
                        <div class="performance-label">1 Hafta</div>
                        <div class="performance-value">{format_perf(perf_7d)}</div>
                    </div>
                </div>
                
                <!-- Technical Metrics -->
                <div class="stats-grid">
                    <div class="metric">
                        <span class="metric-label">RSI</span> 
                        <span class="metric-value">{fund['RSI']}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">MACD</span> 
                        <span class="metric-value">{fund['MACD']}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Fiyat</span> 
                        <span class="metric-value">{fund['Fiyat']} TL</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Yatırımcı</span> 
                        <span class="metric-value">{investor_str} ({fund['Yatırımcı_Değişim']})</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Hacim</span> 
                        <span class="metric-value">{fund['Hacim']} ({fund['Hacim_Değişim']})</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Volatilite</span> 
                        <span class="metric-value">{fund.get('volatility', 'N/A')}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Sharpe</span> 
                        <span class="metric-value">{fund.get('sharpe_ratio', 'N/A')}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">5G Pozitif</span> 
                        <span class="metric-value">{fund.get('positive_days_5', 'N/A')}/5</span>
                    </div>
                </div>
                {f'<div style="margin-top: 12px; font-size: 0.85em; color: #6c757d;"><strong>Kriterler:</strong> {fund["Kriterler"]}</div>' if fund.get('Kriterler') and fund['Kriterler'] != "Belirgin trend sinyali yok" else ""}
            </div>
"""
        
        portfolio_html += """
        </div>
"""
        return portfolio_html

    def _generate_candidates_section(self, early_momentum: List[CandidateResult], trend_candidates: List[CandidateResult]) -> str:
        """Generate candidates section with performance sorting and top 10 filter"""
        candidates_html = ""
        
        # Sort early momentum by weekly performance and take top 10
        sorted_early_momentum = sorted(early_momentum, 
                                     key=lambda x: x.fund_statistics.get('performance_7d', 0), 
                                     reverse=True)[:10]  # Limit to top 10 by weekly performance
        
        # Early Momentum Candidates
        candidates_html += f"""
        <div class="section">
            <h2>🟡 Erken Momentum Adayları</h2>
            <div style="background: rgba(243, 156, 18, 0.1); padding: 12px; border-radius: 10px; margin-bottom: 20px; font-size: 0.85em;">
                📊 Haftalık performansa göre en iyi 10 aday gösteriliyor (toplam: {len(early_momentum)} aday)
            </div>
"""
        
        if sorted_early_momentum:
            for candidate in sorted_early_momentum:
                # Extract fund statistics and performance
                investor_count = candidate.fund_statistics.get('investor_count', 'N/A')
                investor_change = candidate.fund_statistics.get('investor_change_week', 'N/A')
                volume = candidate.fund_statistics.get('volume', 'N/A')
                volume_change = candidate.fund_statistics.get('volume_change_week', 'N/A')
                
                perf_1d = candidate.fund_statistics.get('performance_1d', 0)
                perf_3d = candidate.fund_statistics.get('performance_3d', 0) 
                perf_7d = candidate.fund_statistics.get('performance_7d', 0)
                
                # Format performance with colors
                def format_perf(value):
                    if value > 0:
                        return f"<span class='perf-positive'>+{value:.2f}%</span>"
                    elif value < 0:
                        return f"<span class='perf-negative'>{value:.2f}%</span>"
                    else:
                        return f"<span class='perf-neutral'>0.00%</span>"
                
                candidates_html += f"""
            <div class="candidate-item early-momentum">
                <h3>{candidate.fund_code} 
                    <span style="background: linear-gradient(135deg, #666, #888); color: white; padding: 2px 8px; border-radius: 12px; font-size: 0.7em; font-weight: normal; margin-left: 8px;">🏢 {candidate.company}</span>
                    <span style="background: linear-gradient(135deg, #f39c12, #e67e22); color: white; padding: 3px 10px; border-radius: 15px; font-size: 0.75em; font-weight: bold; margin-left: 8px;">🟡 SKOR: {candidate.score:.0f}</span>
                    <a href="https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={candidate.fund_code}" 
                       target="_blank" class="tefas-btn">TEFAS'ta Gör</a>
                </h3>
                
                <!-- Performance Metrics -->
                <div class="performance-grid">
                    <div class="performance-metric">
                        <div class="performance-label">1 Gün</div>
                        <div class="performance-value">{format_perf(perf_1d)}</div>
                    </div>
                    <div class="performance-metric">
                        <div class="performance-label">3 Gün</div>
                        <div class="performance-value">{format_perf(perf_3d)}</div>
                    </div>
                    <div class="performance-metric">
                        <div class="performance-label">1 Hafta</div>
                        <div class="performance-value">{format_perf(perf_7d)}</div>
                    </div>
                </div>
                
                <!-- Technical and Fund Metrics -->
                <div class="stats-grid">
                    <div class="metric">
                        <span class="metric-label">RSI</span> 
                        <span class="metric-value">{candidate.rsi}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">MACD</span> 
                        <span class="metric-value">{candidate.macd}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Fiyat</span> 
                        <span class="metric-value">{candidate.price:.4f} TL</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Yatırımcı</span> 
                        <span class="metric-value">{investor_count} ({investor_change})</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Hacim</span> 
                        <span class="metric-value">{volume} ({volume_change})</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Volatilite</span> 
                        <span class="metric-value">{candidate.fund_statistics.get('volatility', 'N/A')}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Sharpe</span> 
                        <span class="metric-value">{candidate.fund_statistics.get('sharpe_ratio', 'N/A')}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">5G Pozitif</span> 
                        <span class="metric-value">{candidate.fund_statistics.get('positive_days_5', 'N/A')}/5</span>
                    </div>
                </div>
                <div style="margin-top: 12px; font-size: 0.85em; color: #6c757d;"><strong>Kriterler:</strong> {candidate.criteria}</div>
            </div>
"""
        else:
            candidates_html += """
            <p>🔍 Şu anda erken momentum adayı yok</p>
"""
        
        candidates_html += """
        </div>
"""
        
        # Sort trend candidates by weekly performance and take top 10
        sorted_trend_candidates = sorted(trend_candidates, 
                                       key=lambda x: x.fund_statistics.get('performance_7d', 0), 
                                       reverse=True)[:10]  # Limit to top 10 by weekly performance
        
        # Trend Candidates
        candidates_html += f"""
        <div class="section">
            <h2>📎 Güçlü Trend Adayları</h2>
            <div style="background: rgba(39, 174, 96, 0.1); padding: 12px; border-radius: 10px; margin-bottom: 20px; font-size: 0.85em;">
                📊 Haftalık performansa göre en iyi 10 aday gösteriliyor (toplam: {len(trend_candidates)} aday)
            </div>
"""
        
        if sorted_trend_candidates:
            for candidate in sorted_trend_candidates:
                # Extract fund statistics and performance
                investor_count = candidate.fund_statistics.get('investor_count', 'N/A')
                investor_change = candidate.fund_statistics.get('investor_change_week', 'N/A')
                volume = candidate.fund_statistics.get('volume', 'N/A')
                volume_change = candidate.fund_statistics.get('volume_change_week', 'N/A')
                
                perf_1d = candidate.fund_statistics.get('performance_1d', 0)
                perf_3d = candidate.fund_statistics.get('performance_3d', 0) 
                perf_7d = candidate.fund_statistics.get('performance_7d', 0)
                
                # Format performance with colors
                def format_perf(value):
                    if value > 0:
                        return f"<span class='perf-positive'>+{value:.2f}%</span>"
                    elif value < 0:
                        return f"<span class='perf-negative'>{value:.2f}%</span>"
                    else:
                        return f"<span class='perf-neutral'>0.00%</span>"
                
                candidates_html += f"""
            <div class="candidate-item strong-trend">
                <h3>{candidate.fund_code} 
                    <span style="background: linear-gradient(135deg, #666, #888); color: white; padding: 2px 8px; border-radius: 12px; font-size: 0.7em; font-weight: normal; margin-left: 8px;">🏢 {candidate.company}</span>
                    <span style="background: linear-gradient(135deg, #27ae60, #229954); color: white; padding: 3px 10px; border-radius: 15px; font-size: 0.75em; font-weight: bold; margin-left: 8px;">💎 SKOR: {candidate.score:.0f}</span>
                    <a href="https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={candidate.fund_code}" 
                       target="_blank" class="tefas-btn">TEFAS'ta Gör</a>
                </h3>
                
                <!-- Performance Metrics -->
                <div class="performance-grid">
                    <div class="performance-metric">
                        <div class="performance-label">1 Gün</div>
                        <div class="performance-value">{format_perf(perf_1d)}</div>
                    </div>
                    <div class="performance-metric">
                        <div class="performance-label">3 Gün</div>
                        <div class="performance-value">{format_perf(perf_3d)}</div>
                    </div>
                    <div class="performance-metric">
                        <div class="performance-label">1 Hafta</div>
                        <div class="performance-value">{format_perf(perf_7d)}</div>
                    </div>
                </div>
                
                <!-- Technical and Fund Metrics -->
                <div class="stats-grid">
                    <div class="metric">
                        <span class="metric-label">RSI</span> 
                        <span class="metric-value">{candidate.rsi}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">MACD</span> 
                        <span class="metric-value">{candidate.macd}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Fiyat</span> 
                        <span class="metric-value">{candidate.price:.4f} TL</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Yatırımcı</span> 
                        <span class="metric-value">{investor_count} ({investor_change})</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Hacim</span> 
                        <span class="metric-value">{volume} ({volume_change})</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Volatilite</span> 
                        <span class="metric-value">{candidate.fund_statistics.get('volatility', 'N/A')}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">Sharpe</span> 
                        <span class="metric-value">{candidate.fund_statistics.get('sharpe_ratio', 'N/A')}</span>
                    </div>
                    <div class="metric">
                        <span class="metric-label">5G Pozitif</span> 
                        <span class="metric-value">{candidate.fund_statistics.get('positive_days_5', 'N/A')}/5</span>
                    </div>
                </div>
                <div style="margin-top: 12px; font-size: 0.85em; color: #6c757d;"><strong>Kriterler:</strong> {candidate.criteria}</div>
            </div>
"""
        else:
            candidates_html += """
            <p>🔍 Şu anda güçlü trend adayı yok</p>
"""
        
        candidates_html += """
        </div>
"""
        
        return candidates_html

    def _generate_alerts_section(self, declining_funds: List[Dict[str, Any]]) -> str:
        """Generate alerts section"""
        alerts_html = """
        <div class="section">
            <h2>⚠️ Portföy Uyarıları</h2>
"""
        
        if declining_funds:
            alerts_html += f"""
            <p><strong>🚨 {len(declining_funds)} fonda düşüş sinyali tespit edildi!</strong></p>
"""
            for fund in declining_funds:
                alerts_html += f"""
            <div class="candidate-item decline-warning">
                <h3>🔴 {fund['Fon']} 
                    <span style="color: #6c757d;">| Skor: {fund['Skor']}</span>
                    <a href="https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={fund['Fon']}" 
                       target="_blank" class="tefas-btn">TEFAS'ta Gör</a>
                </h3>
                <p><strong>Kriterler:</strong> {fund['Kriterler']}</p>
            </div>
"""
        else:
            alerts_html += """
            <p>✅ Şu anda portföyde düşüş uyarısı yok</p>
"""
        
        alerts_html += """
        </div>
"""
        return alerts_html

    def _generate_case_results_section(self, case_results: Dict[str, List]) -> str:
        """Generate case results section with improved formatting"""
        case_titles = {
            "case1": "① Son 3 İşlem Günü Art Arda ≥ %1.0",
            "case2": "② Son 3 İşlem Günü Toplam ≥ %4.0",
            "case3": "③ Son 5 İşlem Günü Toplam ≥ %5.0"
        }
        
        case_html = """
        <div class="section">
            <h2>📊 Case Analiz Sonuçları</h2>
"""
        
        # Show summary if no cases found
        total_cases = sum(len(results) for results in case_results.values())
        if total_cases == 0:
            case_html += """
            <p style="text-align: center; color: #6c757d; padding: 40px; font-style: italic;">
                🔍 Bu analiz periyodunda özel pattern tespit edilmedi.<br>
                <small>System 64 fondan sample aldı ve case kriterlerini kontrol etti.</small>
            </p>
"""
        else:
            case_html += f"""
            <div style="background: rgba(102, 126, 234, 0.1); padding: 15px; border-radius: 12px; margin-bottom: 25px; border-left: 4px solid #667eea;">
                <strong>📈 Toplam {total_cases} özel pattern tespit edildi</strong><br>
                <small style="opacity: 0.8;">64 fondan alınan sample üzerinde analiz yapıldı</small>
            </div>
"""
        
        for case_type, results in case_results.items():
            if results:
                # Sort results by cumulative percentage (highest to lowest) and take top 15
                sorted_results = sorted(results, 
                                      key=lambda x: x.get('cumulative_pct', 0), 
                                      reverse=True)[:15]  # Limit to top 15 by weekly performance
                
                total_results = len(results)
                title = case_titles.get(case_type, f"Case {case_type}")
                case_html += f"""
            <div class="case-section">
                <h3 class="case-title">{title}</h3>
                <div style="background: rgba(243, 156, 18, 0.05); padding: 8px 12px; border-radius: 8px; margin-bottom: 15px; font-size: 0.8em; color: #6c757d;">
                    📈 Haftalık performansa göre en iyi 15 fon gösteriliyor (toplam: {total_results} fon)
                </div>
"""
                
                for result in sorted_results:
                    fund_code = result.get('fund', result.get('Fon', 'N/A'))
                    
                    # Format dates better
                    dates = result.get('dates', result.get('Tarihler', []))
                    formatted_dates = []
                    for date_str in dates:
                        if isinstance(date_str, str) and len(date_str) >= 10:
                            # Convert 2025-10-13 to 13.10
                            parts = date_str.split('-')
                            if len(parts) >= 3:
                                formatted_dates.append(f"{parts[2]}.{parts[1]}")
                            else:
                                formatted_dates.append(date_str)
                        else:
                            formatted_dates.append(str(date_str))
                    
                    # Format daily percentages
                    daily_pct = result.get('daily_pct', result.get('Günlük %', []))
                    if isinstance(daily_pct, list):
                        formatted_daily = []
                        for pct in daily_pct:
                            if pct is not None:
                                if pct >= 0:
                                    formatted_daily.append(f"+{pct:.2f}%")
                                else:
                                    formatted_daily.append(f"{pct:.2f}%")
                            else:
                                formatted_daily.append("N/A")
                        daily_display = " | ".join(formatted_daily)
                    else:
                        daily_display = str(daily_pct)
                    
                    # Get cumulative percentage
                    cum_pct = result.get('cumulative_pct', result.get('Kümülatif % (3gün)', result.get('Kümülatif % (5gün)', 'N/A')))
                    if isinstance(cum_pct, (int, float)):
                        cum_display = f"+{cum_pct:.2f}%" if cum_pct >= 0 else f"{cum_pct:.2f}%"
                    else:
                        cum_display = str(cum_pct)
                    
                    # Get enhanced data
                    company = result.get('company', 'Bilinmeyen')
                    score = result.get('score', 'N/A')
                    rsi = result.get('rsi', 'N/A')
                    macd = result.get('macd', 'N/A')
                    price = result.get('price', 'N/A')
                    criteria = result.get('criteria', 'Pattern sinyali')
                    investor_count = result.get('investor_count', 'N/A')
                    investor_change = result.get('investor_change_week', 'N/A')
                    volume = result.get('volume', 'N/A')
                    volume_change = result.get('volume_change_week', 'N/A')
                    perf_1d = result.get('performance_1d', 0)
                    perf_3d = result.get('performance_3d', 0)
                    perf_7d = result.get('performance_7d', 0)
                    
                    # Format performance
                    def format_perf(value):
                        if isinstance(value, (int, float)):
                            if value > 0:
                                return f"<span class='perf-positive'>+{value:.2f}%</span>"
                            elif value < 0:
                                return f"<span class='perf-negative'>{value:.2f}%</span>"
                            else:
                                return f"<span class='perf-neutral'>0.00%</span>"
                        return str(value)
                    
                    case_html += f"""
                <div class="candidate-item" style="border-left-color: #f39c12;">
                    <h3>{fund_code} 
                        <span style="background: linear-gradient(135deg, #666, #888); color: white; padding: 2px 8px; border-radius: 12px; font-size: 0.7em; font-weight: normal; margin-left: 8px;">🏢 {company}</span>
                        <span style="background: linear-gradient(135deg, #f39c12, #e67e22); color: white; padding: 3px 10px; border-radius: 15px; font-size: 0.75em; font-weight: bold; margin-left: 8px;">📊 SKOR: {score}</span>
                        <a href="https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={fund_code}" 
                           target="_blank" class="tefas-btn">TEFAS'ta Gör</a>
                        <span style="background: linear-gradient(135deg, #27ae60, #2ecc71); color: white; padding: 3px 10px; border-radius: 15px; font-size: 0.75em; font-weight: bold; margin-left: 8px;">📈 Kümülatif: {cum_display}</span>
                    </h3>
                    
                    <!-- Performance Metrics -->
                    <div class="performance-grid">
                        <div class="performance-metric">
                            <div class="performance-label">1 Gün</div>
                            <div class="performance-value">{format_perf(perf_1d)}</div>
                        </div>
                        <div class="performance-metric">
                            <div class="performance-label">3 Gün</div>
                            <div class="performance-value">{format_perf(perf_3d)}</div>
                        </div>
                        <div class="performance-metric">
                            <div class="performance-label">1 Hafta</div>
                            <div class="performance-value">{format_perf(perf_7d)}</div>
                        </div>
                    </div>
                    
                    <!-- Technical and Fund Metrics -->
                    <div class="stats-grid">
                        <div class="metric">
                            <span class="metric-label">RSI</span> 
                            <span class="metric-value">{rsi}</span>
                        </div>
                        <div class="metric">
                            <span class="metric-label">MACD</span> 
                            <span class="metric-value">{macd}</span>
                        </div>
                        <div class="metric">
                            <span class="metric-label">Fiyat</span> 
                            <span class="metric-value">{price:.4f} TL</span>
                        </div>
                        <div class="metric">
                            <span class="metric-label">Yatırımcı</span> 
                            <span class="metric-value">{investor_count} ({investor_change})</span>
                        </div>
                        <div class="metric">
                            <span class="metric-label">Hacim</span> 
                            <span class="metric-value">{volume} ({volume_change})</span>
                        </div>
                        <div class="metric">
                            <span class="metric-label">Volatilite</span> 
                            <span class="metric-value">{result.get('volatility', 'N/A')}</span>
                        </div>
                        <div class="metric">
                            <span class="metric-label">Sharpe</span> 
                            <span class="metric-value">{result.get('sharpe_ratio', 'N/A')}</span>
                        </div>
                        <div class="metric">
                            <span class="metric-label">5 Gün Pozitif Kapanış</span> 
                            <span class="metric-value">{result.get('positive_days_5', 0)}/5 gün</span>
                        </div>
                    </div>
                    
                    <div style="display: grid; grid-template-columns: 1fr; gap: 12px; margin: 12px 0;">
                        <div class="metric">
                            <span class="metric-label">📅 Tarihler</span> 
                            <span class="metric-value" style="font-size: 0.85em; white-space: nowrap; overflow: hidden; text-overflow: ellipsis;">{' → '.join(formatted_dates)}</span>
                        </div>
                    </div>
                    
                    <div style="background: rgba(0,0,0,0.05); padding: 8px 10px; border-radius: 6px; font-size: 0.8em; margin-bottom: 12px;">
                        <strong style="color: #495057;">📊 Günlük:</strong> 
                        <span style="font-family: 'SF Mono', 'Monaco', 'Menlo', monospace; color: #2c3e50; white-space: nowrap;">{daily_display}</span>
                    </div>
                    
                    <div style="margin-top: 12px; font-size: 0.85em; color: #6c757d;"><strong>Kriterler:</strong> {criteria}</div>
                </div>
"""
                
                case_html += """
            </div>
"""
        
        case_html += """
        </div>
"""
        return case_html

    def _generate_statistics_section(self, stats: Any) -> str:
        """Generate statistics section"""
        if not stats:
            return ""
            
        return f"""
        <div class="section">
            <h2>📊 Analiz İstatistikleri</h2>
            <div class="stats-grid">
                <div class="stat-item">
                    <div class="metric-label">Toplam Süre</div>
                    <div class="metric-value">{stats.elapsed_time/60:.1f} dakika</div>
                </div>
                <div class="stat-item">
                    <div class="metric-label">Toplam Fonlar</div>
                    <div class="metric-value">{stats.total_funds:,}</div>
                </div>
                <div class="stat-item">
                    <div class="metric-label">Analiz Edilen</div>
                    <div class="metric-value">{stats.processed_funds:,}</div>
                </div>
                <div class="stat-item">
                    <div class="metric-label">Filtrelenen</div>
                    <div class="metric-value">{stats.filtered_out:,}</div>
                </div>
                <div class="stat-item">
                    <div class="metric-label">İşlem Hızı</div>
                    <div class="metric-value">{stats.processing_rate:.1f} fon/saniye</div>
                </div>
                <div class="stat-item">
                    <div class="metric-label">Erken Momentum</div>
                    <div class="metric-value">{stats.early_momentum_count} aday</div>
                </div>
                <div class="stat-item">
                    <div class="metric-label">Güçlü Trend</div>
                    <div class="metric-value">{stats.trend_candidates_count} aday</div>
                </div>
                <div class="stat-item">
                    <div class="metric-label">Portföy Uyarısı</div>
                    <div class="metric-value">{stats.portfolio_alerts} fon</div>
                </div>
            </div>
        </div>
"""

    def _generate_html_footer(self) -> str:
        """Generate HTML footer"""
        return """
        <div class="section" style="text-align: center; background: #f8f9fa;">
            <p style="margin: 0; color: #6c757d;">
                🤖 Bu rapor TEFAS OOP v2.0 sistemi tarafından otomatik olarak oluşturulmuştur.<br>
                📅 Her iş günü otomatik olarak güncellenir.
            </p>
        </div>
    </div>
</body>
</html>
"""

    def generate_excel_report(
        self, 
        panel_data: Dict[str, Any],
        portfolio_analysis: List[Dict[str, Any]],
        early_momentum: List[CandidateResult], 
        trend_candidates: List[CandidateResult],
        declining_funds: List[Dict[str, Any]],
        case_results: Dict[str, List],
        stats: Any
    ) -> Optional[str]:
        """Generate comprehensive Excel report with all analysis data"""
        if not EXCEL_AVAILABLE:
            print("⚠️ Excel raporu için openpyxl gerekli: pip install openpyxl")
            return None
        
        try:
            print("📊 Excel raporu oluşturuluyor...")
            
            # Create workbook and style definitions
            wb = Workbook()
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4472C4")
            subheader_font = Font(bold=True, color="2F5597")
            center_alignment = Alignment(horizontal="center", vertical="center")
            wrap_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            
            # Color definitions
            green_fill = PatternFill("solid", fgColor="E8F5E8")
            red_fill = PatternFill("solid", fgColor="FFEBEE")
            yellow_fill = PatternFill("solid", fgColor="FFF9C4")
            gold_fill = PatternFill("solid", fgColor="FFF8DC")
            
            # 1. SUMMARY SHEET
            ws_summary = wb.active
            ws_summary.title = "📊 Özet"
            
            self._create_summary_sheet(ws_summary, panel_data, declining_funds, early_momentum, trend_candidates, stats)
            
            # 2. PORTFOLIO SHEET
            if portfolio_analysis:
                ws_portfolio = wb.create_sheet(title="💼 Portföy")
                self._create_portfolio_sheet(ws_portfolio, portfolio_analysis)
            
            # 3. EARLY MOMENTUM SHEET
            if early_momentum:
                ws_early = wb.create_sheet(title="🚀 Erken Momentum")
                self._create_candidates_sheet(ws_early, early_momentum, "Erken Momentum")
            
            # 4. TREND CANDIDATES SHEET
            if trend_candidates:
                ws_trend = wb.create_sheet(title="💪 Trend Adayları")
                self._create_candidates_sheet(ws_trend, trend_candidates, "Trend Adayları")
            
            # 5. CASE ANALYSIS SHEETS (separate tabs)
            if case_results.get('case1'):
                ws_case1 = wb.create_sheet(title="① 3Gün Art Arda")
                self._create_case_sheet(ws_case1, case_results['case1'], "Son 3 İşlem Günü Art Arda ≥ %1.0")
            
            if case_results.get('case2'):
                ws_case2 = wb.create_sheet(title="② 3Gün Toplam")
                self._create_case_sheet(ws_case2, case_results['case2'], "Son 3 İşlem Günü Toplam ≥ %4.0")
            
            if case_results.get('case3'):
                ws_case3 = wb.create_sheet(title="③ 5Gün Toplam")
                self._create_case_sheet(ws_case3, case_results['case3'], "Son 5 İşlem Günü Toplam ≥ %5.0")
            
            # 6. ALERTS SHEET
            if declining_funds:
                ws_alerts = wb.create_sheet(title="⚠️ Uyarılar")
                self._create_alerts_sheet(ws_alerts, declining_funds)
            
            # Save Excel file
            filename = f"tefas_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            wb.save(filename)
            
            print(f"✅ Excel raporu kaydedildi: {filename}")
            print(f"📂 Excel dosyasını açmak için: open {filename}")
            
            return filename
            
        except Exception as e:
            print(f"❌ Excel raporu oluşturma hatası: {e}")
            return None
    
    def _create_summary_sheet(self, ws, panel_data, declining_funds, early_momentum, trend_candidates, stats):
        """Create summary sheet with overview statistics"""
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"TEFAS Günlük Takip Paneli - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A1'].font = Font(bold=True, size=16, color="2F5597")
        ws['A1'].alignment = center_alignment
        
        # Summary data
        summary_data = [
            ["Kategori", "Adet", "Durum"],
            ["💼 Portföy Fonları", len(panel_data.get('portfolio_funds', [])), "Aktif"],
            ["🚀 Erken Momentum", len(early_momentum), "İzlemede"],
            ["💪 Güçlü Trend", len(trend_candidates), "İzlemede"],
            ["⚠️ Aktif Uyarılar", len(declining_funds), "Dikkat"]
        ]
        
        for row_idx, row_data in enumerate(summary_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx == 3:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                else:
                    cell.alignment = center_alignment
        
        # Statistics section
        if stats:
            row = 9
            ws[f'A{row}'] = "📊 Analiz İstatistikleri:"
            ws[f'A{row}'].font = Font(bold=True)
            
            stats_data = [
                f"⏱️ Süre: {stats.elapsed_time/60:.1f} dakika",
                f"📋 Toplam Fonlar: {stats.total_funds:,}",
                f"✅ İşlenen: {stats.processed_funds:,}", 
                f"🔽 Filtrelenen: {stats.filtered_out:,}",
                f"⚡ Hız: {stats.processing_rate:.1f} fon/saniye"
            ]
            
            for i, stat in enumerate(stats_data):
                ws[f'B{row+i}'] = stat
                ws[f'B{row+i}'].font = Font(size=10)
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
    
    def _create_portfolio_sheet(self, ws, portfolio_analysis):
        """Create portfolio analysis sheet"""
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_alignment = Alignment(horizontal="center", vertical="center")
        wrap_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        green_fill = PatternFill("solid", fgColor="E8F5E8")
        red_fill = PatternFill("solid", fgColor="FFEBEE")
        yellow_fill = PatternFill("solid", fgColor="FFF9C4")
        
        # Headers
        headers = ["Fon Kodu", "Şirket", "Trend Durumu", "RSI", "MACD", "Fiyat (TL)", 
                  "Yatırımcı", "Yatırımcı Değişim", "Hacim", "Hacim Değişim", "Kriterler"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, fund in enumerate(portfolio_analysis, 2):
            data = [
                fund['Fon'],
                fund.get('Şirket', 'Bilinmeyen'),
                fund['Trend_Status'],
                fund['RSI'],
                fund['MACD'],
                fund['Fiyat'],
                fund['Yatırımcı'],
                fund['Yatırımcı_Değişim'],
                fund['Hacim'],
                fund['Hacim_Değişim'],
                fund['Kriterler']
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Color coding based on trend status
                if col_idx == 3:  # Trend Status column (moved to position 3)
                    if "🔥" in str(value) or "📈" in str(value):
                        cell.fill = green_fill
                    elif "🔴" in str(value) or "📉" in str(value):
                        cell.fill = red_fill
                    else:
                        cell.fill = yellow_fill
                
                # Alignment
                if col_idx in [4, 5, 6]:  # RSI, MACD, Price (adjusted for new column)
                    cell.alignment = center_alignment
                elif col_idx == 11:  # Criteria (adjusted for new column)
                    cell.alignment = wrap_alignment
        
        # Column widths (adjusted for new Company column)
        column_widths = [12, 18, 20, 8, 12, 12, 12, 15, 15, 20, 40]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width
    
    def _create_candidates_sheet(self, ws, candidates, sheet_type):
        """Create candidates sheet (early momentum or trend)"""
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        center_alignment = Alignment(horizontal="center", vertical="center")
        wrap_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        headers = ["Fon Kodu", "Şirket", "Skor", "RSI", "MACD", "Fiyat (TL)", "Yatırımcı", "Yatırımcı Değişim", "Hacim", "Hacim Değişim", "Kriterler"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        for row_idx, candidate in enumerate(candidates, 2):
            # Extract fund statistics
            investor_count = candidate.fund_statistics.get('investor_count', 'N/A')
            investor_change = candidate.fund_statistics.get('investor_change_week', 'N/A')
            volume = candidate.fund_statistics.get('volume', 'N/A')
            volume_change = candidate.fund_statistics.get('volume_change_week', 'N/A')
            
            data = [
                candidate.fund_code,
                candidate.company,
                candidate.score,
                candidate.rsi,
                candidate.macd,
                f"{candidate.price:.4f}",
                investor_count,
                investor_change,
                volume,
                volume_change,
                candidate.criteria
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx == 11:  # Criteria (moved to column 11)
                    cell.alignment = wrap_alignment
                elif col_idx in [3, 4, 5, 6]:  # Score, RSI, MACD, Price (adjusted for new column)
                    cell.alignment = center_alignment
        
        # Column widths (adjusted for new Company column)
        column_widths = [12, 18, 8, 10, 12, 12, 12, 15, 15, 20, 40]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + idx)].width = width
    
    def _create_case_sheet(self, ws, case_data, case_title):
        """Create individual case analysis sheet"""
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4472C4")
        subheader_font = Font(bold=True, color="2F5597")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        
        # Title
        ws['A1'] = case_title
        ws['A1'].font = Font(bold=True, size=14, color="2F5597")
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ["Fon Kodu", "Tarihler", "Günlük %", "Kümülatif %", "TEFAS Linki"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, result in enumerate(case_data, 4):
            fund_code = result.get('fund', result.get('Fon', 'N/A'))
            dates = ', '.join(result.get('dates', result.get('Tarihler', [])))
            daily_pct = ', '.join([f"%{x:.1f}" for x in result.get('daily_pct', result.get('Günlük %', []))])
            cumulative_pct = f"%{result.get('cumulative_pct', result.get('Kümülatif % (3gün)', result.get('Kümülatif % (5gün)', 0)))}"
            tefas_link = f"https://www.tefas.gov.tr/FonAnaliz.aspx?FonKod={fund_code}"
            
            data = [fund_code, dates, daily_pct, cumulative_pct, tefas_link]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx == 5:  # TEFAS link column
                    cell.value = "TEFAS'ta Gör"
                    cell.hyperlink = tefas_link
                    cell.font = Font(color="0000FF", underline="single")
                    cell.alignment = center_alignment
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _create_alerts_sheet(self, ws, declining_funds):
        """Create alerts/warnings sheet"""
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="DC3545")
        center_alignment = Alignment(horizontal="center", vertical="center")
        wrap_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        red_fill = PatternFill("solid", fgColor="FFEBEE")
        
        headers = ["Fon Kodu", "Uyarı Türü", "Skor", "Kriterler"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        for row_idx, fund in enumerate(declining_funds, 2):
            data = [
                fund.get('Fon', 'Bilinmiyor'),
                "Düşüş Sinyali",
                fund.get('Skor', 'N/A'),
                fund.get('Kriterler', 'Detay yok')
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.fill = red_fill
                if col_idx == 4:  # Criteria
                    cell.alignment = wrap_alignment
                else:
                    cell.alignment = center_alignment
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 40


if __name__ == "__main__":
    # Test report generator
    report_gen = ReportGenerator()
    print("📊 ReportGenerator sınıfı hazır!")
    print(f"Excel desteği: {'✅' if EXCEL_AVAILABLE else '❌'}")
