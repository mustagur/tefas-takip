#!/usr/bin/env python3
"""
TEFAS Backtest Engine

Decision report sinyallerini T+1 fiyatlamayla replay eder.
İki mod: SAFE_MODE (AL) ve AGGRESSIVE_MODE (AL + ALINABİLİR ADAY)
Çıktı: backtest_report.xlsx (2 sheet)

Kritik kural:
  - Giriş fiyatı = sinyal gününden sonraki ilk iş günü NAV'ı
  - Çıkış fiyatı = sinyal gününden sonraki ilk iş günü NAV'ı
  - Aynı fonda aynı anda max 1 açık pozisyon
"""

import json
import glob
import os
import sys
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from data_provider import FundDataProvider

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")
REPORT_FILE = os.path.join(BASE_DIR, "backtest_report.xlsx")

# Sinyal tanımları
ENTRY_SIGNALS_SAFE = {"AL"}
ENTRY_SIGNALS_AGGRESSIVE = {"AL", "ALINABİLİR ADAY"}
EXIT_SIGNALS = {"SAT", "AZALT"}

# Aggressive mod günlük yeni pozisyon limiti
AGGRESSIVE_DAILY_LIMIT = 3


class BacktestEngine:
    """T+1 fiyatlama ile backtest motoru"""

    def __init__(self):
        self.data_provider = FundDataProvider()
        self.price_cache = {}  # fund_code -> DataFrame

    # ------------------------------------------------------------------
    # Veri yükleme
    # ------------------------------------------------------------------

    def load_decision_reports(self):
        """outputs/ altındaki tüm decision_report JSON dosyalarını yükle"""
        reports = {}
        pattern = os.path.join(OUTPUT_DIR, "decision_report_*.json")
        for filepath in sorted(glob.glob(pattern)):
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    data = json.load(f)
                date_str = data.get("date")
                decisions = data.get("decisions", [])
                if date_str and decisions:
                    reports[date_str] = decisions
            except (json.JSONDecodeError, KeyError) as e:
                print(f"  ⚠️ Atlanıyor {os.path.basename(filepath)}: {e}")
        return reports

    # ------------------------------------------------------------------
    # Fiyat lookup (T+1 kuralı)
    # ------------------------------------------------------------------

    def _ensure_price_data(self, fund_code, start_date, end_date):
        """Fon fiyat verisini çek ve cache'le"""
        if fund_code not in self.price_cache:
            df = self.data_provider.get_fund_data(fund_code, start_date, end_date)
            if df is not None and not df.empty:
                self.price_cache[fund_code] = df
            else:
                self.price_cache[fund_code] = pd.DataFrame()
        return self.price_cache[fund_code]

    def get_next_available_price(self, fund_code, signal_date_str, start_date, end_date):
        """
        T+1 kuralı: signal_date'den SONRAKİ ilk iş gününün NAV'ını döndür.
        Hafta sonu ve tatiller doğal olarak atlanır (veri olmayan günler).

        Returns: (date_str, price) veya (None, None)
        """
        df = self._ensure_price_data(fund_code, start_date, end_date)
        if df.empty:
            return None, None

        signal_date = pd.Timestamp(signal_date_str)
        future = df[df["date"] > signal_date]

        if future.empty:
            return None, None

        row = future.iloc[0]
        return row["date"].strftime("%Y-%m-%d"), round(float(row["price"]), 4)

    def get_latest_price(self, fund_code, start_date, end_date):
        """Mevcut en güncel NAV fiyatını döndür (açık pozisyonlar için)"""
        df = self._ensure_price_data(fund_code, start_date, end_date)
        if df.empty:
            return None, None

        row = df.iloc[-1]
        return row["date"].strftime("%Y-%m-%d"), round(float(row["price"]), 4)

    # ------------------------------------------------------------------
    # Daily features (ret_5d lookup için)
    # ------------------------------------------------------------------

    def _load_daily_features(self, date_str):
        """daily_features JSON'ından ret_5d lookup dict'ı döndür"""
        filepath = os.path.join(OUTPUT_DIR, f"daily_features_{date_str}.json")
        if not os.path.exists(filepath):
            return {}
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            return {fund["code"]: fund.get("ret_5d", 0) or 0 for fund in data.get("funds", [])}
        except (json.JSONDecodeError, KeyError):
            return {}

    # ------------------------------------------------------------------
    # Backtest çalıştır
    # ------------------------------------------------------------------

    def run(self, mode="safe"):
        """
        Backtest çalıştır.

        Args:
            mode: "safe" (sadece AL) veya "aggressive" (AL + ALINABİLİR ADAY)

        Returns:
            List[dict] — trade listesi
        """
        reports = self.load_decision_reports()
        if not reports:
            print("  ⚠️ Hiç decision report bulunamadı")
            return []

        entry_signals = ENTRY_SIGNALS_SAFE if mode == "safe" else ENTRY_SIGNALS_AGGRESSIVE

        # Fiyat verisi çekilecek tarih aralığı
        all_dates = sorted(reports.keys())
        start_dt = datetime.strptime(all_dates[0], "%Y-%m-%d") - timedelta(days=7)
        end_dt = datetime.now() + timedelta(days=7)
        start_date = start_dt.strftime("%Y-%m-%d")
        end_date = end_dt.strftime("%Y-%m-%d")

        # Sinyal içeren fonları bul (sadece bunların fiyatı çekilecek)
        relevant_funds = set()
        for decisions in reports.values():
            for d in decisions:
                if d["decision"] in entry_signals or d["decision"] in EXIT_SIGNALS:
                    relevant_funds.add(d["code"])

        print(f"  📊 {len(reports)} gün rapor, {len(relevant_funds)} fon fiyatı çekilecek")

        trades = []
        open_positions = {}  # fund_code -> trade dict referansı (trades listesindeki)

        for date_str in sorted(reports.keys()):
            decisions = reports[date_str]

            # ========== FAZ 1: ÇIKIŞLAR (limit yok) ==========
            for decision in decisions:
                code = decision["code"]
                signal = decision["decision"]

                if code not in relevant_funds:
                    continue

                if code in open_positions and signal in EXIT_SIGNALS:
                    trade = open_positions[code]
                    exit_date, exit_price = self.get_next_available_price(
                        code, date_str, start_date, end_date
                    )

                    trade["exit_signal_date"] = date_str
                    trade["exit_date"] = exit_date
                    trade["exit_price"] = exit_price
                    trade["exit_reason"] = signal
                    trade["status"] = "CLOSED"

                    if trade["entry_price"] and exit_price:
                        trade["return_pct"] = round(
                            (exit_price - trade["entry_price"])
                            / trade["entry_price"]
                            * 100,
                            2,
                        )
                    if trade["entry_date"] and exit_date:
                        trade["days_held"] = (
                            datetime.strptime(exit_date, "%Y-%m-%d")
                            - datetime.strptime(trade["entry_date"], "%Y-%m-%d")
                        ).days

                    del open_positions[code]

            # ========== FAZ 2: GİRİŞLER ==========
            if mode == "aggressive":
                # Önce AL sinyalleri (yüksek öncelik)
                al_candidates = [
                    d for d in decisions
                    if d["decision"] == "AL"
                    and d["code"] not in open_positions
                    and d["code"] in relevant_funds
                ]
                al_candidates.sort(key=lambda x: (-x.get("score_final", 0), x.get("rsi", 100)))

                # Sonra ALINABİLİR (score_final desc, rsi asc, ret_5d asc)
                ret_5d_lookup = self._load_daily_features(date_str)
                alinabilir = [
                    d for d in decisions
                    if d["decision"] == "ALINABİLİR ADAY"
                    and d["code"] not in open_positions
                    and d["code"] in relevant_funds
                ]
                alinabilir.sort(key=lambda x: (
                    -x.get("score_final", 0),
                    x.get("rsi", 100),
                    ret_5d_lookup.get(x["code"], 0),
                ))

                # AL önce, ardından ALINABİLİR — günlük max 3
                ordered_candidates = al_candidates + alinabilir
                daily_opens = 0

                for decision in ordered_candidates:
                    if daily_opens >= AGGRESSIVE_DAILY_LIMIT:
                        break

                    code = decision["code"]
                    entry_date, entry_price = self.get_next_available_price(
                        code, date_str, start_date, end_date
                    )
                    if entry_date is None or entry_price is None:
                        continue

                    trade = {
                        "fund": code,
                        "signal_date": date_str,
                        "entry_date": entry_date,
                        "exit_signal_date": None,
                        "exit_date": None,
                        "entry_price": entry_price,
                        "exit_price": None,
                        "return_pct": None,
                        "days_held": None,
                        "status": "OPEN",
                        "exit_reason": None,
                    }
                    open_positions[code] = trade
                    trades.append(trade)
                    daily_opens += 1

            else:
                # Safe mod: sadece AL, limit yok
                for decision in decisions:
                    code = decision["code"]
                    signal = decision["decision"]

                    if code not in relevant_funds:
                        continue
                    if code in open_positions or signal not in entry_signals:
                        continue

                    entry_date, entry_price = self.get_next_available_price(
                        code, date_str, start_date, end_date
                    )
                    if entry_date is None or entry_price is None:
                        continue

                    trade = {
                        "fund": code,
                        "signal_date": date_str,
                        "entry_date": entry_date,
                        "exit_signal_date": None,
                        "exit_date": None,
                        "entry_price": entry_price,
                        "exit_price": None,
                        "return_pct": None,
                        "days_held": None,
                        "status": "OPEN",
                        "exit_reason": None,
                    }
                    open_positions[code] = trade
                    trades.append(trade)

        # Açık pozisyonlar için unrealized P&L hesapla
        for code, trade in open_positions.items():
            latest_date, latest_price = self.get_latest_price(
                code, start_date, end_date
            )
            if latest_price and trade["entry_price"]:
                trade["exit_price"] = latest_price
                trade["exit_date"] = latest_date
                trade["return_pct"] = round(
                    (latest_price - trade["entry_price"])
                    / trade["entry_price"]
                    * 100,
                    2,
                )
                trade["days_held"] = (
                    datetime.strptime(latest_date, "%Y-%m-%d")
                    - datetime.strptime(trade["entry_date"], "%Y-%m-%d")
                ).days

        return trades

    # ------------------------------------------------------------------
    # Summary metrikler
    # ------------------------------------------------------------------

    @staticmethod
    def calculate_summary(trades):
        """Trade listesinden özet metrikler hesapla"""
        if not trades:
            return {
                "total_trades": 0,
                "open_trades": 0,
                "closed_trades": 0,
                "win_rate": 0.0,
                "avg_gain": 0.0,
                "avg_loss": 0.0,
                "avg_return": 0.0,
                "best_trade": 0.0,
                "worst_trade": 0.0,
                "total_return": 0.0,
            }

        closed = [
            t for t in trades if t["status"] == "CLOSED" and t["return_pct"] is not None
        ]
        open_t = [t for t in trades if t["status"] == "OPEN"]
        wins = [t for t in closed if t["return_pct"] > 0]
        losses = [t for t in closed if t["return_pct"] <= 0]
        all_returns = [t["return_pct"] for t in trades if t["return_pct"] is not None]

        return {
            "total_trades": len(trades),
            "open_trades": len(open_t),
            "closed_trades": len(closed),
            "win_rate": round(len(wins) / len(closed) * 100, 1) if closed else 0.0,
            "avg_gain": round(
                sum(t["return_pct"] for t in wins) / len(wins), 2
            )
            if wins
            else 0.0,
            "avg_loss": round(
                sum(t["return_pct"] for t in losses) / len(losses), 2
            )
            if losses
            else 0.0,
            "avg_return": round(sum(all_returns) / len(all_returns), 2)
            if all_returns
            else 0.0,
            "best_trade": round(max(all_returns), 2) if all_returns else 0.0,
            "worst_trade": round(min(all_returns), 2) if all_returns else 0.0,
            "total_return": round(
                sum(t["return_pct"] for t in closed), 2
            )
            if closed
            else 0.0,
        }


# ======================================================================
# Excel oluşturma
# ======================================================================

# Renkler
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_GREEN = Font(color="006100")
_RED = Font(color="9C0006")
_BLUE = Font(color="0070C0", bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_DATE_FONT = Font(italic=True, color="888888")
_METRIC_LABEL = Font(bold=True, size=10)
_METRIC_VALUE = Font(size=10)

_TRADE_HEADERS = [
    "Fund",
    "Signal Date",
    "Entry Date",
    "Exit Signal Date",
    "Exit Date",
    "Entry Price",
    "Exit Price",
    "Return %",
    "Days Held",
    "Status",
    "Exit Reason",
]


def _write_sheet(ws, title, trades, summary):
    """Bir sheet'e summary + trade tablosu yaz"""
    ws.title = title

    # Başlık
    ws.cell(row=1, column=1, value=f"TEFAS Backtest — {title}").font = _TITLE_FONT
    ws.cell(
        row=2,
        column=1,
        value=f"Son Güncelleme: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ).font = _DATE_FONT

    # Summary metrikleri (2 satır x 5 kolon düzeni)
    metrics = [
        ("Total Trades", summary["total_trades"]),
        ("Open Trades", summary["open_trades"]),
        ("Closed Trades", summary["closed_trades"]),
        ("Win Rate", f"{summary['win_rate']}%"),
        ("Total Return", f"{summary['total_return']}%"),
        ("Avg Return", f"{summary['avg_return']}%"),
        ("Avg Gain", f"{summary['avg_gain']}%"),
        ("Avg Loss", f"{summary['avg_loss']}%"),
        ("Best Trade", f"{summary['best_trade']}%"),
        ("Worst Trade", f"{summary['worst_trade']}%"),
    ]

    base_row = 4
    for i, (label, value) in enumerate(metrics):
        col = (i % 5) * 2 + 1
        row = base_row + i // 5
        ws.cell(row=row, column=col, value=label).font = _METRIC_LABEL
        ws.cell(row=row, column=col + 1, value=value).font = _METRIC_VALUE

    # Trade tablo başlıkları
    table_row = base_row + 3
    for col_idx, header in enumerate(_TRADE_HEADERS, 1):
        cell = ws.cell(row=table_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Trade verileri
    for i, trade in enumerate(trades):
        r = table_row + 1 + i
        ws.cell(row=r, column=1, value=trade["fund"])
        ws.cell(row=r, column=2, value=trade["signal_date"])
        ws.cell(row=r, column=3, value=trade["entry_date"])
        ws.cell(row=r, column=4, value=trade["exit_signal_date"] or "")
        ws.cell(row=r, column=5, value=trade["exit_date"] or "")
        ws.cell(row=r, column=6, value=trade["entry_price"])
        ws.cell(row=r, column=7, value=trade["exit_price"])

        ret = trade["return_pct"]
        ret_cell = ws.cell(row=r, column=8, value=ret)
        if ret is not None:
            ret_cell.number_format = "0.00"
            ret_cell.font = _GREEN if ret > 0 else _RED

        ws.cell(row=r, column=9, value=trade["days_held"])

        status_cell = ws.cell(row=r, column=10, value=trade["status"])
        if trade["status"] == "OPEN":
            status_cell.font = _BLUE

        ws.cell(row=r, column=11, value=trade["exit_reason"] or "")

    # Kolon genişlikleri
    widths = [8, 13, 13, 16, 13, 12, 12, 10, 10, 10, 12]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def generate_excel(safe_trades, aggressive_trades, safe_summary, agg_summary):
    """backtest_report.xlsx oluştur (2 sheet)"""
    wb = Workbook()

    # Sheet 1: SAFE_MODE
    _write_sheet(wb.active, "SAFE_MODE", safe_trades, safe_summary)

    # Sheet 2: AGGRESSIVE_MODE
    ws2 = wb.create_sheet("AGGRESSIVE_MODE")
    _write_sheet(ws2, "AGGRESSIVE_MODE", aggressive_trades, agg_summary)

    wb.save(REPORT_FILE)
    print(f"\n📊 Backtest raporu: {REPORT_FILE}")


# ======================================================================
# Ana çalıştırıcı
# ======================================================================


def main():
    print("=" * 50)
    print("📊 TEFAS BACKTEST ENGINE")
    print("=" * 50)

    engine = BacktestEngine()

    # SAFE_MODE
    print("\n🔒 SAFE_MODE backtesting...")
    safe_trades = engine.run(mode="safe")
    safe_summary = engine.calculate_summary(safe_trades)
    print(
        f"   Trades: {safe_summary['total_trades']} | "
        f"Open: {safe_summary['open_trades']} | "
        f"Win Rate: {safe_summary['win_rate']}% | "
        f"Total Return: {safe_summary['total_return']}%"
    )

    # AGGRESSIVE_MODE
    print("\n🚀 AGGRESSIVE_MODE backtesting...")
    aggressive_trades = engine.run(mode="aggressive")
    agg_summary = engine.calculate_summary(aggressive_trades)
    print(
        f"   Trades: {agg_summary['total_trades']} | "
        f"Open: {agg_summary['open_trades']} | "
        f"Win Rate: {agg_summary['win_rate']}% | "
        f"Total Return: {agg_summary['total_return']}%"
    )

    # Excel oluştur
    generate_excel(safe_trades, aggressive_trades, safe_summary, agg_summary)

    print("\n✅ Backtest tamamlandı!")


if __name__ == "__main__":
    main()
